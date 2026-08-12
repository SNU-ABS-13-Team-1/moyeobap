#!/usr/bin/env python3
"""통합 조사표 엑셀을 앱용 CSV 정본으로 변환한다.

    python3 scripts/xlsx_to_csv.py

입력  시흥캠퍼스_식당카페_조사표_통합.xlsx  (팀이 수집·검수하는 작업 파일)
출력  data/restaurants.csv, data/menus.csv   (앱이 참조하는 정본)

정본은 CSV다. 엑셀은 수집 단계의 작업 파일이므로, 엑셀을 고친 뒤에는 이
스크립트를 다시 돌려 CSV를 갱신한다. CSV를 직접 고쳤다면 엑셀에도 같은
내용을 반영해야 다음 실행에서 되돌아가지 않는다.

sources.csv는 생성하지 않는다. 엑셀에 출처·확인일 열이 없어 값을 지어내게
되기 때문이다. DATA_GUIDE 4절을 보고 직접 작성한다.
"""

import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl이 필요합니다: pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "시흥캠퍼스_식당카페_조사표_통합.xlsx"
SHEET = "매장·메뉴 입력"

# 엑셀 열 순서. 조사표 서식이 바뀌면 여기부터 고친다.
COL = {
    "no": 0, "name": 1, "major": 2, "minor": 3, "address": 4,
    "hours": 5, "closed": 6, "phone": 7, "rating": 8,
    "delivery": 9, "min_order": 10, "menu_no": 11,
    "menu_name": 12, "menu_price": 13,
}

MAJOR_TO_TYPE = {"식당": "restaurant", "카페": "cafe"}

RESTAURANT_FIELDS = [
    "restaurant_id", "name", "type", "category", "min_order", "delivery_time",
    "address", "phone", "business_hours", "closed_days", "rating", "image_url",
]
MENU_FIELDS = ["menu_id", "restaurant_id", "name", "price", "image_url"]


def clean(value):
    """빈 셀은 빈 문자열로. 값을 지어내지 않는다."""
    if value is None:
        return ""
    return str(value).strip()


def to_int(value, where):
    text = clean(value).replace(",", "").replace("원", "")
    if not text:
        return ""
    try:
        return str(int(float(text)))
    except ValueError:
        print(f"  경고: {where} 숫자로 읽을 수 없음 -> {value!r}", file=sys.stderr)
        return ""


def main():
    if not XLSX.exists():
        sys.exit(f"조사표를 찾을 수 없습니다: {XLSX}")

    workbook = openpyxl.load_workbook(XLSX, data_only=True)
    if SHEET not in workbook.sheetnames:
        sys.exit(f"'{SHEET}' 시트가 없습니다. 시트: {workbook.sheetnames}")

    rows = [
        row for row in workbook[SHEET].iter_rows(min_row=2, values_only=True)
        if row and row[COL["no"]] is not None
    ]

    restaurants, menus, seen_menu_ids = {}, [], set()

    for row in rows:
        store_no = clean(row[COL["no"]])
        store_id = f"store-{store_no}"
        major = clean(row[COL["major"]])

        if store_id not in restaurants:
            store_type = MAJOR_TO_TYPE.get(major, "")
            if not store_type:
                print(f"  경고: {store_id} 대분류가 '식당'/'카페'가 아님 -> {major!r}",
                      file=sys.stderr)
            restaurants[store_id] = {
                "restaurant_id": store_id,
                "name": clean(row[COL["name"]]),
                "type": store_type,
                "category": clean(row[COL["minor"]]),
                "min_order": to_int(row[COL["min_order"]], f"{store_id} 최소주문금액"),
                "delivery_time": clean(row[COL["delivery"]]),
                "address": clean(row[COL["address"]]),
                "phone": clean(row[COL["phone"]]),
                "business_hours": clean(row[COL["hours"]]),
                "closed_days": clean(row[COL["closed"]]),
                "rating": clean(row[COL["rating"]]),
                "image_url": "",  # 사용 권한 확인 전까지 비워 둔다 (DATA_GUIDE 5절)
            }

        menu_name = clean(row[COL["menu_name"]])
        if not menu_name:
            continue

        menu_no = clean(row[COL["menu_no"]]) or str(len(menus) + 1)
        menu_id = f"{store_id}-menu-{menu_no}"
        if menu_id in seen_menu_ids:
            print(f"  경고: 메뉴 ID 중복 -> {menu_id} ({menu_name})", file=sys.stderr)
            continue
        seen_menu_ids.add(menu_id)

        menus.append({
            "menu_id": menu_id,
            "restaurant_id": store_id,
            "name": menu_name,
            "price": to_int(row[COL["menu_price"]], f"{menu_id} 가격"),
            "image_url": "",
        })

    write_csv(ROOT / "data/restaurants.csv", RESTAURANT_FIELDS, restaurants.values())
    write_csv(ROOT / "data/menus.csv", MENU_FIELDS, menus)

    print(f"매장 {len(restaurants)}곳 -> data/restaurants.csv")
    print(f"메뉴 {len(menus)}개 -> data/menus.csv")
    print("sources.csv는 생성하지 않았습니다. DATA_GUIDE 4절을 보고 직접 작성하세요.")


def write_csv(path, fields, records):
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(records)


if __name__ == "__main__":
    main()
